import logging
from openpyxl import Workbook
from tqdm import tqdm
from random import randint
from time import sleep as pause
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_price_page(driver, url):
    """Парсит страницу товара и возвращает цену."""
    try:
        driver.get(url)
        pause(randint(7, 11))

        # Увеличиваем время ожидания до 30 секунд
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, "product-buy__price"))
        )

        soup = BeautifulSoup(driver.page_source, 'lxml')
        price = soup.find('div', class_="product-buy__price")

        # Обработка цены
        if price:
            price_text = price.text.strip().replace('&nbsp;', ' ').replace('₽', '').replace(' ', '')
            price_value = int(price_text) if price_text.isdigit() else 0
            logging.info(f"Цена для {url}: {price_value} рублей")
        else:
            logging.warning(f"Цена не найдена для {url}")
            price_value = 0

        return price_value

    except Exception as e:
        logging.error(f"Ошибка при парсинге страницы {url}: {e}")
        return None


def get_urls_from_page(driver):
    """Собирает все ссылки на текущей странице."""
    WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "catalog-product__name"))
    )
    soup = BeautifulSoup(driver.page_source, 'lxml')
    elements = soup.find_all('a', class_="catalog-product__name ui-link ui-link_black")
    urls = ['https://www.dns-shop.ru' + element.get("href") for element in elements]
    logging.info(f"Найдено {len(urls)} товаров на текущей странице.")
    return urls


def to_excel(prices, urls, file_name="prices"):
    workbook = Workbook()
    sheet = workbook.active
    logging.info('Начался экспорт в Excel Таблицу')

    # Заголовки
    sheet.cell(row=1, column=1, value="Цена")
    sheet.cell(row=1, column=2, value="Ссылка")

    for index, (price, url) in enumerate(zip(prices, urls), 2):
        sheet.cell(row=index, column=1, value=price)
        sheet.cell(row=index, column=2, value=url)

    workbook.save(f"{file_name}.xlsx")


def main():
    driver = uc.Chrome()
    url_to_parse = 'https://www.dns-shop.ru/catalog/17a89bb916404e77/platy-rasshireniya/?p={page}'

    prices = []
    urls = []  # Список для хранения ссылок на товары
    page_number = 1

    try:
        while True:
            # Получаем ссылки на товары на текущей странице
            driver.get(url_to_parse.format(page=page_number))
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, "catalog-product__name"))
            )
            page_urls = get_urls_from_page(driver)

            for url in tqdm(page_urls, ncols=70, unit='товаров', colour='blue'):
                price = parse_price_page(driver, url)
                if price is not None:  # Проверяем, что цена была успешно получена
                    prices.append(price)
                    urls.append(url)  # Добавляем ссылку на товар

            # Проверяем, есть ли следующая страница
            try:
                next_page = driver.find_element(By.CSS_SELECTOR, '.pagination-widget__page:not(.pagination-widget__page_disabled)')
                if next_page:
                    page_number += 1  # Увеличиваем номер страницы, если следующая страница доступна
                else:
                    break  # Если элемента нет, выходим из цикла
            except NoSuchElementException:
                logging.warning("Кнопка следующей страницы не найдена.")
                break
            except Exception as e:
                logging.warning("Произошла ошибка: " + str(e))
                break

        to_excel(prices, urls, file_name="prices")  # Передаем ссылки в функцию экспорта

    except Exception as e:
        logging.error(f"Произошла ошибка: {e}")

    finally:
        driver.quit()


if __name__ == '__main__':
    main()
    logging.info('Все готово!')